import asyncio
import csv
import hashlib
import io
import json
import re
from pathlib import Path
from uuid import UUID, uuid4

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field
from pypdf import PdfReader
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError

from app.core.auth import AuthenticatedUser, get_current_user
from app.core.config import Settings, get_settings
from app.core.knowledge_access import require_knowledge_base_permission
from app.db.session import get_db_connection
from app.services.embeddings import embed_texts, vector_literal
from app.services.storage import (
    LocalKnowledgeStorage,
    StorageConfigurationError,
    StorageError,
    get_knowledge_storage,
    get_knowledge_storage_for_provider,
)

router = APIRouter(prefix="/knowledge-bases", tags=["knowledge"])

SUPPORTED_EXTENSIONS = {
    ".csv": "text/csv",
    ".json": "application/json",
    ".md": "text/markdown",
    ".pdf": "application/pdf",
    ".txt": "text/plain",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
SAFE_FILENAME_PATTERN = re.compile(r"[^a-zA-Z0-9._-]+")
CHUNK_SIZE = 1200
CHUNK_OVERLAP = 120


class KnowledgeFileSummary(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    byte_size: int = Field(alias="byteSize")
    created_at: str = Field(alias="createdAt")
    error_message: str | None = Field(default=None, alias="errorMessage")
    file_hash: str = Field(alias="fileHash")
    file_id: str = Field(alias="fileId")
    knowledge_base_id: str = Field(alias="knowledgeBaseId")
    mime_type: str = Field(alias="mimeType")
    original_name: str = Field(alias="originalName")
    status: str
    storage_provider: str = Field(alias="storageProvider")
    updated_at: str = Field(alias="updatedAt")
    workspace_id: str = Field(alias="workspaceId")


class KnowledgeFileListResponse(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    files: list[KnowledgeFileSummary]


FILE_SELECT = text(
    """
    SELECT
        "byteSize" AS byte_size,
        "createdAt" AS created_at,
        "errorMessage" AS error_message,
        "fileHash" AS file_hash,
        "id" AS file_id,
        "knowledgeBaseId" AS knowledge_base_id,
        "mimeType" AS mime_type,
        "originalName" AS original_name,
        "status" AS status,
        "storageProvider" AS storage_provider,
        "storageKey" AS storage_key,
        "updatedAt" AS updated_at,
        "workspaceId" AS workspace_id
    FROM "KnowledgeFile"
    """
)

FILE_BY_ID_QUERY = text(
    FILE_SELECT.text
    + """
    WHERE "id" = :file_id
      AND "knowledgeBaseId" = :knowledge_base_id
      AND "workspaceId" = :workspace_id
    LIMIT 1
    """
)

FILE_PROCESS_QUERY = text(
    FILE_SELECT.text
    + """
    WHERE "id" = :file_id
      AND "workspaceId" = :workspace_id
    LIMIT 1
    """
)

FILE_LIST_QUERY = text(
    FILE_SELECT.text
    + """
    WHERE "knowledgeBaseId" = :knowledge_base_id
      AND "workspaceId" = :workspace_id
    ORDER BY "createdAt" DESC
    """
)

FILE_INSERT_QUERY = text(
    """
    INSERT INTO "KnowledgeFile"
        (
            "byteSize", "fileHash", "id", "knowledgeBaseId", "mimeType",
            "originalName", "storageKey", "storageProvider", "uploadedBy", "workspaceId"
        )
    VALUES
        (
            :byte_size, :file_hash, :file_id, :knowledge_base_id, :mime_type,
            :original_name, :storage_key, :storage_provider, :uploaded_by, :workspace_id
        )
    ON CONFLICT ("knowledgeBaseId", "fileHash") DO NOTHING
    RETURNING "id"
    """
)

FILE_STATUS_QUERY = text(
    """
    UPDATE "KnowledgeFile"
    SET "errorMessage" = :error_message,
        "status" = :status,
        "updatedAt" = CURRENT_TIMESTAMP
    WHERE "id" = :file_id
      AND "workspaceId" = :workspace_id
    """
)

CHUNKS_DELETE_QUERY = text('DELETE FROM "KnowledgeChunk" WHERE "fileId" = :file_id')
CHUNKS_INSERT_QUERY = text(
    """
    INSERT INTO "KnowledgeChunk"
        ("chunkIndex", "content", "fileId", "knowledgeBaseId", "metadata", "workspaceId")
    VALUES
        (
            :chunk_index,
            :content,
            :file_id,
            :knowledge_base_id,
            CAST(:metadata AS jsonb),
            :workspace_id
        )
    """
)

CHUNKS_INSERT_WITH_EMBEDDING_QUERY = text(
    """
    INSERT INTO "KnowledgeChunk"
        (
            "chunkIndex", "content", "embedding", "fileId", "knowledgeBaseId",
            "metadata", "workspaceId"
        )
    VALUES
        (
            :chunk_index,
            :content,
            CAST(:embedding AS vector),
            :file_id,
            :knowledge_base_id,
            CAST(:metadata AS jsonb),
            :workspace_id
        )
    """
)

FILE_DELETE_QUERY = text(
    """
    DELETE FROM "KnowledgeFile"
    WHERE "id" = :file_id
      AND "knowledgeBaseId" = :knowledge_base_id
      AND "workspaceId" = :workspace_id
    RETURNING "storageKey", "storageProvider"
    """
)


def _iso_timestamp(value: object) -> str:
    from datetime import UTC, datetime

    if not isinstance(value, datetime):
        return str(value)
    timestamp = value if value.tzinfo else value.replace(tzinfo=UTC)
    return timestamp.isoformat(timespec="milliseconds").replace("+00:00", "Z")


def _file_summary(row: dict[str, object]) -> KnowledgeFileSummary:
    return KnowledgeFileSummary(
        byteSize=int(row["byte_size"]),
        createdAt=_iso_timestamp(row["created_at"]),
        errorMessage=row["error_message"] if isinstance(row["error_message"], str) else None,
        fileHash=str(row["file_hash"]),
        fileId=str(row["file_id"]),
        knowledgeBaseId=str(row["knowledge_base_id"]),
        mimeType=str(row["mime_type"]),
        originalName=str(row["original_name"]),
        status=str(row["status"]),
        storageProvider=str(row["storage_provider"]),
        updatedAt=_iso_timestamp(row["updated_at"]),
        workspaceId=str(row["workspace_id"]),
    )


def _database_error(message: str) -> JSONResponse:
    return JSONResponse(
        status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
        content={"code": "database:unavailable", "cause": message},
    )


def _feature_disabled() -> JSONResponse:
    return JSONResponse(
        status_code=status.HTTP_409_CONFLICT,
        content={
            "code": "knowledge_ingestion:disabled",
            "message": "Knowledge ingestion is disabled until its migration is applied.",
        },
    )


def _safe_filename(filename: str) -> str:
    candidate = SAFE_FILENAME_PATTERN.sub("_", filename).strip("._")
    return (candidate or "upload")[:160]


def _extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def _storage_path(settings: Settings, storage_key: str) -> Path:
    storage = get_knowledge_storage(settings)
    if not isinstance(storage, LocalKnowledgeStorage):
        raise HTTPException(
            status_code=400,
            detail="Storage paths are only available for local knowledge storage.",
        )
    try:
        return storage.path_for(storage_key)
    except StorageError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error


async def _best_effort_delete(settings: Settings, storage_key: str) -> None:
    try:
        await get_knowledge_storage(settings).delete(storage_key)
    except (StorageConfigurationError, StorageError):
        return


def _extract_text(filename: str, content: bytes) -> str:
    extension = _extension(filename)
    if extension in {".txt", ".md", ".json"}:
        return content.decode("utf-8-sig", errors="replace")
    if extension == ".csv":
        rows = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")))
        return "\n".join("\t".join(cell.strip() for cell in row) for row in rows)
    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            lines.append(f"[Sheet: {worksheet.title}]")
            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    lines.append("\t".join(values))
        workbook.close()
        return "\n".join(lines)
    if extension == ".pdf":
        reader = PdfReader(io.BytesIO(content))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    raise ValueError("Unsupported knowledge file type.")


def _chunk_text(content: str) -> list[str]:
    normalized = content.replace("\r\n", "\n").strip()
    if not normalized:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(normalized):
        end = min(start + CHUNK_SIZE, len(normalized))
        chunk = normalized[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end == len(normalized):
            break
        start = max(end - CHUNK_OVERLAP, start + 1)
    return chunks


async def process_knowledge_file(file_id: UUID, workspace_id: UUID) -> None:
    settings = get_settings()
    try:
        async with get_db_connection() as connection:
            result = await connection.execute(
                FILE_PROCESS_QUERY,
                {
                    "file_id": file_id,
                    "workspace_id": workspace_id,
                },
            )
            row = result.mappings().first()
    except (RuntimeError, SQLAlchemyError):
        return

    if row is None:
        return

    try:
        async with get_db_connection() as connection:
            async with connection.begin():
                await connection.execute(
                    FILE_STATUS_QUERY,
                    {
                        "error_message": None,
                        "file_id": file_id,
                        "status": "processing",
                        "workspace_id": workspace_id,
                    },
                )

        storage = get_knowledge_storage_for_provider(
            settings,
            str(row["storage_provider"]),
        )
        content = await storage.read(str(row["storage_key"]))
        extracted_text = await asyncio.to_thread(
            _extract_text,
            str(row["original_name"]),
            content,
        )
        chunks = _chunk_text(extracted_text)
        if not chunks:
            raise ValueError("No extractable text was found in the uploaded file.")
        embeddings = (
            await embed_texts(chunks, settings)
            if settings.knowledge_embeddings_enabled
            else None
        )

        async with get_db_connection() as connection:
            async with connection.begin():
                await connection.execute(CHUNKS_DELETE_QUERY, {"file_id": file_id})
                await connection.execute(
                    CHUNKS_INSERT_WITH_EMBEDDING_QUERY
                    if embeddings is not None
                    else CHUNKS_INSERT_QUERY,
                    [
                        {
                            "chunk_index": index,
                            "content": chunk,
                            "embedding": vector_literal(embeddings[index])
                            if embeddings is not None
                            else None,
                            "file_id": file_id,
                            "knowledge_base_id": row["knowledge_base_id"],
                            "metadata": json.dumps(
                                {
                                    "fileName": row["original_name"],
                                    "chunkIndex": index,
                                },
                                separators=(",", ":"),
                            ),
                            "workspace_id": workspace_id,
                        }
                        for index, chunk in enumerate(chunks)
                    ],
                )
                await connection.execute(
                    FILE_STATUS_QUERY,
                    {
                        "error_message": None,
                        "file_id": file_id,
                        "status": "ready",
                        "workspace_id": workspace_id,
                    },
                )
    except Exception as error:
        try:
            async with get_db_connection() as connection:
                async with connection.begin():
                    await connection.execute(
                        FILE_STATUS_QUERY,
                        {
                            "error_message": str(error)[:1000],
                            "file_id": file_id,
                            "status": "failed",
                            "workspace_id": workspace_id,
                        },
                    )
        except (RuntimeError, SQLAlchemyError):
            return


@router.get("/{knowledge_base_id}/files", response_model=KnowledgeFileListResponse)
async def list_knowledge_files(
    knowledge_base_id: UUID,
    workspace_id: UUID = Query(..., alias="workspace_id"),
    current_user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> KnowledgeFileListResponse | JSONResponse:
    await require_knowledge_base_permission(
        current_user,
        workspace_id,
        knowledge_base_id,
        "read",
    )
    if not settings.knowledge_ingestion_enabled:
        return _feature_disabled()

    try:
        async with get_db_connection() as connection:
            result = await connection.execute(
                FILE_LIST_QUERY,
                {
                    "knowledge_base_id": knowledge_base_id,
                    "workspace_id": workspace_id,
                },
            )
            rows = result.mappings().all()
    except RuntimeError as error:
        return _database_error(str(error))
    except SQLAlchemyError:
        return _database_error("FastAPI could not list knowledge files.")

    return KnowledgeFileListResponse(files=[_file_summary(dict(row)) for row in rows])


@router.post("/{knowledge_base_id}/files", response_model=None, status_code=202)
async def upload_knowledge_file(
    knowledge_base_id: UUID,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    workspace_id: UUID = Query(..., alias="workspace_id"),
    current_user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict[str, KnowledgeFileSummary] | JSONResponse:
    await require_knowledge_base_permission(
        current_user,
        workspace_id,
        knowledge_base_id,
        "manage",
    )
    if not settings.knowledge_ingestion_enabled:
        return _feature_disabled()

    if current_user.is_development:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="A persisted user is required to upload knowledge files.",
        )
    try:
        uploaded_by = UUID(current_user.user_id)
    except ValueError as error:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="The authenticated user is not linked to a local workspace.",
        ) from error

    original_name = _safe_filename(file.filename or "upload")
    extension = _extension(original_name)
    if extension not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            detail="Supported file types are PDF, CSV, XLSX, JSON, Markdown, and text.",
        )

    content = await file.read(settings.knowledge_max_file_bytes + 1)
    if len(content) > settings.knowledge_max_file_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="The knowledge file is larger than the configured limit.",
        )

    file_id = uuid4()
    storage_key = (
        f"{workspace_id}/{knowledge_base_id}/{file_id}-{original_name}"
    )
    try:
        storage = get_knowledge_storage(settings)
        await storage.put(storage_key, content)
    except StorageConfigurationError as error:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={"code": "storage:misconfigured", "cause": str(error)},
        )
    except StorageError as error:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={"code": "storage:unavailable", "cause": str(error)},
        )

    file_hash = hashlib.sha256(content).hexdigest()
    try:
        async with get_db_connection() as connection:
            async with connection.begin():
                result = await connection.execute(
                    FILE_INSERT_QUERY,
                    {
                        "byte_size": len(content),
                        "file_hash": file_hash,
                        "file_id": file_id,
                        "knowledge_base_id": knowledge_base_id,
                        "mime_type": SUPPORTED_EXTENSIONS[extension],
                        "original_name": original_name,
                        "storage_key": storage_key,
                        "storage_provider": storage.provider,
                        "uploaded_by": uploaded_by,
                        "workspace_id": workspace_id,
                    },
                )
                inserted_id = result.scalar_one_or_none()
                if inserted_id is None:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="This file has already been uploaded to the knowledge base.",
                    )
                file_result = await connection.execute(
                    FILE_BY_ID_QUERY,
                    {
                        "file_id": inserted_id,
                        "knowledge_base_id": knowledge_base_id,
                        "workspace_id": workspace_id,
                    },
                )
                row = file_result.mappings().one()
    except HTTPException:
        await _best_effort_delete(settings, storage_key)
        raise
    except RuntimeError as error:
        await _best_effort_delete(settings, storage_key)
        return _database_error(str(error))
    except SQLAlchemyError:
        await _best_effort_delete(settings, storage_key)
        return _database_error("FastAPI could not create the knowledge file record.")

    background_tasks.add_task(process_knowledge_file, file_id, workspace_id)
    return {"file": _file_summary(dict(row))}


@router.delete("/{knowledge_base_id}/files/{file_id}", response_model=None)
async def delete_knowledge_file(
    knowledge_base_id: UUID,
    file_id: UUID,
    workspace_id: UUID = Query(..., alias="workspace_id"),
    current_user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict[str, bool] | JSONResponse:
    await require_knowledge_base_permission(
        current_user,
        workspace_id,
        knowledge_base_id,
        "manage",
    )
    if not settings.knowledge_ingestion_enabled:
        return _feature_disabled()

    try:
        async with get_db_connection() as connection:
            async with connection.begin():
                result = await connection.execute(
                    FILE_DELETE_QUERY,
                    {
                        "file_id": file_id,
                        "knowledge_base_id": knowledge_base_id,
                        "workspace_id": workspace_id,
                    },
                )
                deleted = result.mappings().first()
                if deleted is None:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail="Knowledge file not found.",
                    )
                storage_key = str(deleted["storageKey"])
                storage_provider = str(deleted["storageProvider"])
    except HTTPException:
        raise
    except RuntimeError as error:
        return _database_error(str(error))
    except SQLAlchemyError:
        return _database_error("FastAPI could not delete the knowledge file.")

    try:
        storage = get_knowledge_storage_for_provider(settings, storage_provider)
        await storage.delete(storage_key)
    except StorageConfigurationError as error:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={"code": "storage:misconfigured", "cause": str(error)},
        )
    except StorageError as error:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={"code": "storage:unavailable", "cause": str(error)},
        )
    return {"deleted": True}
